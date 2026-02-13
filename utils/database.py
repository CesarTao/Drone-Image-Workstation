import streamlit as st
import pandas as pd
import mysql.connector
import os
import time
import openpyxl

from config import DB_CONFIG, COLUMN_MAPPING
from utils.common import format_size

def get_connection():
    return mysql.connector.connect(**DB_CONFIG)

def save_to_db(data_list):  # 储存到数据库
    if not data_list: return 0

    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        incoming_hashes = [item.get('FileHash') for item in data_list if item.get('FileHash')]

        existing_hashes = set()

        if incoming_hashes:
            placeholders_check = ', '.join(['%s'] * len(incoming_hashes))
            sql_check = f"SELECT FileHash FROM drone_photos WHERE FileHash IN ({placeholders_check})"

            cursor.execute(sql_check, tuple(incoming_hashes))

            result = cursor.fetchall()
            for row in result:
                existing_hashes.add(row[0])

        final_data_list = [
            item for item in data_list
            if item.get('FileHash') not in existing_hashes
        ]

        # 如果过滤完发现全是重复的，直接返回
        if not final_data_list:
            # print("本批次所有数据均为重复，跳过。")
            return 0

        keys = list(final_data_list[0].keys())
        cols = ", ".join(f"`{k}`" for k in keys)
        placeholders = ", ".join(["%s"] * len(keys))

        sql = f"INSERT INTO drone_photos ({cols}) VALUES ({placeholders})"

        # 构造 values 列表
        values = [[item.get(k) for k in keys] for item in final_data_list]

        cursor.executemany(sql, values)
        conn.commit()

        return cursor.rowcount

    except Exception as e:
        st.error(f"入库失败: {e}")
        return 0
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

def clear_all_data():
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE drone_photos")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"清空失败: {e}")
        return False

def load_data_from_db():
    conn = mysql.connector.connect(**DB_CONFIG)
    query = "SELECT * FROM drone_photos ORDER BY capture_time DESC"
    df = pd.read_sql(query, conn)
    conn.close()
    return df

def execute_raw_sql(sql_query):  # 执行原始 SQL 语句并返回 DataFrame
    normalized_sql = sql_query.upper().strip()

    # 安全检查，暂未开启
    # forbidden_keywords = ['DROP', 'DELETE', 'UPDATE', 'TRUNCATE', 'ALTER', 'INSERT', 'GRANT']
    # for word in forbidden_keywords:
    #    # 简单的包含检查，防止 delete from ...
    #    if word in normalized_sql:
    #        return None, f"⚠️ 安全警告：为了保护数据，网页端禁止使用 {word} 命令。请使用数据库管理软件进行操作。"

    conn = None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        # 使用 pandas 直接读取，它能自动处理列名和数据类型
        df_result = pd.read_sql(sql_query, conn)
        if 'FileSize' in df_result.columns:
            df_result['FileSize'] = df_result['FileSize'].apply(format_size)

        df_result = df_result.rename(columns=COLUMN_MAPPING)
        return df_result, None
    except Exception as e:
        return None, str(e)
    finally:
        if conn: conn.close()

def sync_dir_tags(file_path):
    """
    解析文件路径，提取目录，存入 file_dir_tags 表
    """
    if not file_path: return
    
    norm_path = os.path.normpath(file_path).replace('\\', '/')
    folder_path = os.path.dirname(norm_path)
    #print(f"-----------------{folder_path}-------------------")

    parts = norm_path.split('/')
    
    dir_parts = parts[:-1] 
    if not dir_parts: return
    

    l1 = dir_parts[1] if len(dir_parts) >= 1 else ""
    l2 = dir_parts[2] if len(dir_parts) >= 2 else ""
    l3 = dir_parts[3] if len(dir_parts) >= 3 else ""
    folder_name = dir_parts[-1]

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        
        # 插入或忽略 (如果已存在则不覆盖标记状态)
        sql = """
        INSERT IGNORE INTO file_dir_tags 
        (folder_name, full_path, dir_level_1, dir_level_2, dir_level_3) 
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (folder_name, folder_path, l1, l2, l3))
        conn.commit()
    except Exception as e:
        print(f"目录同步失败: {e}")
    finally:
        if conn: conn.close()

def update_marks_batch(df_changes, mode):
    if df_changes.empty: return
    
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    affected_files_count = 0
    
    try:
        #print("------------------- 开始执行批量更新 -----------------------")
        # 遍历修改过的行
        for index, row in df_changes.iterrows():
            
            # 获取和清洗
            f_path_file = row.get('full_path') 
            f_note = row.get('mark_note', '')
            raw_color = row.get('tag_color')
            if pd.isna(raw_color) or raw_color in ["⚪ 无", "无", "nan"]:
                db_color = None
            else:
                db_color = raw_color
            if pd.isna(f_note):
                f_note = ""
            else:
                f_note = str(f_note)
            

            if not f_path_file: 
                print("❌ 跳过：找不到 full_path")
                continue

            # 更新 file_dir_tags 表
            sql_self = "UPDATE file_dir_tags SET tag_color=%s, mark_note=%s WHERE full_path=%s"
            cursor.execute(sql_self, (db_color, f_note, f_path_file))

            if mode == 1:
                continue

            #parent_folder = os.path.dirname(f_path_file)
            parent_folder = f_path_file
            clean_parent = parent_folder.replace('/', '\\') # 确保是反斜杠
            search_pattern = clean_parent.replace('\\', '\\\\') + '\\\\%'
            
            #print(f"处理中: {f_path_file}")
            #print(f" >> 提取父目录: {clean_parent}")
            #print(f" >> SQL匹配模式: {search_pattern}")

            sql_sync = """
            UPDATE drone_photos 
            SET mark_note = %s 
            WHERE FullPath LIKE %s
            """
            cursor.execute(sql_sync, (f_note, search_pattern))
            
            row_count = cursor.rowcount
            affected_files_count += row_count
            #print(f" >> 数据库反馈: 更新了 {row_count} 条记录")
            
        conn.commit()
        st.toast(f"✅ 保存成功！并同步更新了 {row_count} 个文件的备注。")
        time.sleep(1)
        
    except Exception as e:
        st.error(f"保存失败: {e}")
        print(f"ERROR DETAILS: {e}")
    finally:
        if conn: conn.close()


def update_color_by_hashes(file_hashes, new_color):
    """
    批量将指定的一组 file_hash 更新为新颜色
    """
    if not file_hashes: return
    
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    try:
        db_color = None if (new_color == "⚪ 无" or pd.isna(new_color)) else new_color
        
        placeholders = ', '.join(['%s'] * len(file_hashes))
        sql = f"UPDATE file_dir_tags SET tag_color = %s WHERE file_hash IN ({placeholders})"
        
        # 参数：颜色 + 哈希列表
        params = [db_color] + file_hashes
        
        cursor.execute(sql, params)
        conn.commit()
        
        st.toast(f"✅ 成功批量更新了 {cursor.rowcount} 条数据！")
    except Exception as e:
        st.error(f"批量更新失败: {e}")
    finally:
        conn.close()

def process_excel_to_db(uploaded_file):
    """
    解析 Excel，计算时间差，提取起止时间，写入数据库
    """

    def parse_time_custom(val):
        s = str(val).strip()
        if s in ['None', '', 'nan'] or val is None: return None
        offset = 0
        if "+1" in s:
            offset = 24
            s = s.replace("+1", "").strip()
        try:
            parts = s.split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            sec = float(parts[2]) if len(parts) > 2 else 0
            return h + (m / 60) + (sec / 3600) + offset
        except:
            return None

    def decimal_to_str(val):
        if val is None: return ""
        total_seconds = int(val * 3600)
        h = total_seconds // 3600
        m = (total_seconds % 3600) // 60
        s = total_seconds % 60
        # 如果超过24小时（跨夜），保留如 25:00:00 的格式以便区分
        return f"{h:02d}:{m:02d}:{s:02d}"

    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    max_row = ws.max_row
    parsed_times = [None] * (max_row + 1)
    raw_texts = [None] * (max_row + 1)

    target_col_idx = 2
    for row in range(2, max_row + 1):
        cell_val = ws.cell(row=row, column=target_col_idx).value
        parsed_times[row] = parse_time_custom(cell_val)
        raw_texts[row] = str(cell_val) if cell_val is not None else ""

    results_to_insert = []
    batch_id = f"{int(time.time())}"
    current_date_str = "Unknown"

    cnt = 0
    valid_count = 0

    for row in range(2, max_row + 1):
        text = raw_texts[row]
        if text == "":
            cell_val = ws.cell(row=row, column=target_col_idx-1).value
            parsed_times[row] = parse_time_custom(cell_val)
            raw_texts[row] = str(cell_val) if cell_val is not None else ""
            text = raw_texts[row]
            current_date_str = text.strip()
            cnt = 0
        elif "2025-" in text:
            current_date_str = text.strip()
            cnt = 0

        # 计算逻辑 (每4行算一次)
        if cnt % 4 == 0 and cnt != 0 and row > 2:
            curr_val = parsed_times[row]
            prev_val = parsed_times[row - 1]

            if curr_val is not None and prev_val is not None:
                diff = round((curr_val - prev_val) * 60, 2)

                start_str = decimal_to_str(prev_val)
                end_str = decimal_to_str(curr_val)

                results_to_insert.append((
                    batch_id,
                    uploaded_file.name,
                    current_date_str,
                    start_str,
                    end_str,
                    diff
                ))
                valid_count += 1
        cnt += 1

    if results_to_insert:
        conn = mysql.connector.connect(**DB_CONFIG)
        if conn:
            cursor = conn.cursor()
            sql = """
            INSERT INTO task_hours 
            (batch_id, source_filename, task_date, start_time, end_time, duration_minutes)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            try:
                cursor.executemany(sql, results_to_insert)
                conn.commit()
                st.toast(f"成功入库 {len(results_to_insert)} 条记录！")
            except Exception as e:
                st.error(f"数据库写入失败: {e}")
            finally:
                cursor.close()
                conn.close()

    return valid_count